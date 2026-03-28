import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.contrib.auth import get_user_model
import openpyxl

from .models import Farmer, Team1Assignment, Team2Assignment, FarmerUploadLog
from .forms import (
    FarmerUploadForm, BulkAssignTeam1Form, BulkAssignTeam2Form,
    Team1UpdateForm, Team2UpdateForm, FarmerFilterForm
)

User = get_user_model()


def role_required_view(user, *roles):
    return user.role in roles


def dashboard_redirect_core(request):
    """Proxy to accounts dashboard_redirect — needed for URL naming."""
    from accounts.views import dashboard_redirect
    return dashboard_redirect(request)


def role_required(*roles):
    """Decorator factory to restrict views by role."""
    def decorator(view_func):
        @login_required
        def wrapper(request, *args, **kwargs):
            if request.user.role not in roles:
                messages.error(request, "You don't have permission to access this page.")
                return redirect('dashboard_redirect')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ─────────────────────────────────────────────
#  ADMIN VIEWS
# ─────────────────────────────────────────────

@role_required('admin')
def admin_dashboard(request):
    """Stage 1 + Stage 2 overview for Vikas."""
    # Total counts
    total_farmers = Farmer.objects.count()
    assigned_team1 = Team1Assignment.objects.count()
    unassigned = total_farmers - assigned_team1

    # Team 1 stats
    t1_completed = Team1Assignment.objects.filter(call_status='completed').count()
    t1_pending = Team1Assignment.objects.filter(call_status='pending').count()
    t1_interested = Team1Assignment.objects.filter(interest_status='interested').count()
    t1_not_interested = Team1Assignment.objects.filter(interest_status='not_interested').count()
    t1_no_response = Team1Assignment.objects.filter(interest_status='no_response').count()

    # Per-agent Team 1 breakdown
    team1_stats = (
        Team1Assignment.objects
        .values('assigned_to__first_name', 'assigned_to__last_name', 'assigned_to__username')
        .annotate(
            total=Count('id'),
            completed=Count('id', filter=Q(call_status='completed')),
            pending=Count('id', filter=Q(call_status='pending')),
            interested=Count('id', filter=Q(interest_status='interested')),
        )
        .order_by('-total')
    )

    # Team 2 stats
    assigned_team2 = Team2Assignment.objects.count()
    t2_contacted = Team2Assignment.objects.filter(followup_status='contacted').count()
    t2_pending = Team2Assignment.objects.filter(followup_status='pending').count()
    t2_confirmed = Team2Assignment.objects.filter(final_response='confirmed').count()
    t2_declined = Team2Assignment.objects.filter(final_response='declined').count()

    team2_stats = (
        Team2Assignment.objects
        .values('assigned_to__first_name', 'assigned_to__last_name', 'assigned_to__username')
        .annotate(
            total=Count('id'),
            contacted=Count('id', filter=Q(followup_status='contacted')),
            pending=Count('id', filter=Q(followup_status='pending')),
            confirmed=Count('id', filter=Q(final_response='confirmed')),
        )
        .order_by('-total')
    )

    context = {
        'total_farmers': total_farmers,
        'assigned_team1': assigned_team1,
        'unassigned': unassigned,
        't1_completed': t1_completed,
        't1_pending': t1_pending,
        't1_interested': t1_interested,
        't1_not_interested': t1_not_interested,
        't1_no_response': t1_no_response,
        'team1_stats': team1_stats,
        'assigned_team2': assigned_team2,
        't2_contacted': t2_contacted,
        't2_pending': t2_pending,
        't2_confirmed': t2_confirmed,
        't2_declined': t2_declined,
        'team2_stats': team2_stats,
    }
    return render(request, 'core/admin_dashboard.html', context)


@role_required('admin')
def upload_farmers(request):
    """Admin uploads CSV/Excel of farmers — also shows full farmer database."""
    form = FarmerUploadForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        f = request.FILES['file']
        filename = f.name
        success = 0
        failed = 0
        errors = []

        try:
            if filename.endswith('.csv'):
                decoded = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(f, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            for i, row in enumerate(rows, start=2):
                try:
                    farmer_id = str(row.get('farmer_id') or row.get('Farmer ID') or '').strip()
                    name = str(row.get('name') or row.get('Name') or '').strip()
                    phone = str(row.get('phone') or row.get('Phone') or '').strip()
                    if not farmer_id or not name:
                        errors.append(f"Row {i}: Missing farmer_id or name")
                        failed += 1
                        continue
                    Farmer.objects.update_or_create(
                        farmer_id=farmer_id,
                        defaults={
                            'name': name,
                            'phone': phone,
                            'village': str(row.get('village') or row.get('Village') or '').strip(),
                            'district': str(row.get('district') or row.get('District') or '').strip(),
                            'state': str(row.get('state') or row.get('State') or '').strip(),
                            'crop': str(row.get('crop') or row.get('Crop') or '').strip(),
                        }
                    )
                    success += 1
                except Exception as e:
                    errors.append(f"Row {i}: {str(e)}")
                    failed += 1

            FarmerUploadLog.objects.create(
                uploaded_by=request.user,
                file_name=filename,
                total_records=success + failed,
                success_records=success,
                failed_records=failed,
                notes='\n'.join(errors[:20]) if errors else ''
            )
            messages.success(request, f"Upload complete: {success} added/updated, {failed} failed.")
            if errors:
                messages.warning(request, f"Errors in {failed} rows. Check upload log.")
        except Exception as e:
            messages.error(request, f"File processing error: {str(e)}")
        return redirect('upload_farmers')

    # Farmer database view
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    farmers_qs = Farmer.objects.select_related('team1_assignment').all().order_by('farmer_id')

    if search:
        farmers_qs = farmers_qs.filter(
            Q(name__icontains=search) | Q(phone__icontains=search) |
            Q(village__icontains=search) | Q(crop__icontains=search) |
            Q(district__icontains=search)
        )
    if status_filter == 'unassigned':
        farmers_qs = farmers_qs.filter(team1_assignment__isnull=True)
    elif status_filter == 'assigned':
        farmers_qs = farmers_qs.filter(team1_assignment__isnull=False)
    elif status_filter == 'interested':
        farmers_qs = farmers_qs.filter(team1_assignment__interest_status='interested')
    elif status_filter == 'called':
        farmers_qs = farmers_qs.filter(team1_assignment__call_status='completed')

    total_farmers    = Farmer.objects.count()
    assigned_count   = Team1Assignment.objects.count()
    unassigned_count = total_farmers - assigned_count
    interested_count = Team1Assignment.objects.filter(interest_status='interested').count()

    return render(request, 'core/upload_farmers.html', {
        'form': form,
        'farmers': farmers_qs,
        'search': search,
        'total_farmers': total_farmers,
        'assigned_count': assigned_count,
        'unassigned_count': unassigned_count,
        'interested_count': interested_count,
    })


@role_required('admin')
def assign_team1(request):
    """Admin assigns/unassigns farmers to Team 1 — shows ALL farmers in one table."""
    if request.method == 'POST':
        form = BulkAssignTeam1Form(request.POST)
        if form.is_valid():
            ids = [int(x) for x in form.cleaned_data['farmer_ids'].split(',') if x.strip()]
            agent = form.cleaned_data['assigned_to']
            count = 0
            for fid in ids:
                try:
                    farmer = Farmer.objects.get(pk=fid)
                    Team1Assignment.objects.get_or_create(
                        farmer=farmer,
                        defaults={'assigned_to': agent, 'assigned_by': request.user}
                    )
                    count += 1
                except Farmer.DoesNotExist:
                    pass
            messages.success(request, f"{count} farmers assigned to {agent.get_full_name() or agent.username}.")
            return redirect('assign_team1')
    else:
        form = BulkAssignTeam1Form()

    # ALL farmers with assignment info
    all_farmers = Farmer.objects.select_related(
        'team1_assignment__assigned_to'
    ).all().order_by('farmer_id')

    team1_users = User.objects.filter(role='team1')
    team1_summary = []
    for u in team1_users:
        assignments = Team1Assignment.objects.filter(assigned_to=u).select_related('farmer')
        team1_summary.append({
            'user': u,
            'total': assignments.count(),
            'completed': assignments.filter(call_status='completed').count(),
            'pending': assignments.filter(call_status='pending').count(),
            'interested': assignments.filter(interest_status='interested').count(),
            'assignments': assignments,
        })

    return render(request, 'core/assign_team1.html', {
        'all_farmers': all_farmers,
        'farmers': Farmer.objects.filter(team1_assignment__isnull=True),
        'form': form,
        'team1_summary': team1_summary,
        'team1_users': team1_users,
        'total_unassigned': Farmer.objects.filter(team1_assignment__isnull=True).count(),
        'total_assigned': Team1Assignment.objects.count(),
    })


@role_required('admin')
def user_create_team1(request):
    """Quick create a Team 1 member from the assign page modal."""
    if request.method == 'POST':
        username  = request.POST.get('username', '').strip()
        password  = request.POST.get('password', '').strip()
        first_name= request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone     = request.POST.get('phone', '').strip()
        if username and password:
            if User.objects.filter(username=username).exists():
                messages.error(request, f"Username '{username}' already exists. Choose a different username.")
            else:
                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    role='team1'
                )
                user.phone = phone
                user.save()
                messages.success(request, f"Team 1 member '{first_name} {last_name}' added successfully! They can login with username: {username}")
        else:
            messages.error(request, "Username and password are required.")
    return redirect('assign_team1')


@role_required('admin')
def assign_team2(request):
    """Admin assigns/unassigns interested farmers to Team 2 — shows ALL interested farmers."""
    if request.method == 'POST':
        form = BulkAssignTeam2Form(request.POST)
        if form.is_valid():
            ids = [int(x) for x in form.cleaned_data['farmer_ids'].split(',') if x.strip()]
            agent = form.cleaned_data['assigned_to']
            count = 0
            for t1_id in ids:
                try:
                    t1 = Team1Assignment.objects.get(pk=t1_id)
                    Team2Assignment.objects.get_or_create(
                        farmer=t1.farmer,
                        defaults={
                            'team1_assignment': t1,
                            'assigned_to': agent,
                            'assigned_by': request.user
                        }
                    )
                    count += 1
                except Team1Assignment.DoesNotExist:
                    pass
            messages.success(request, f"{count} interested farmers assigned to {agent.get_full_name() or agent.username}.")
            return redirect('assign_team2')
    else:
        form = BulkAssignTeam2Form()

    # ALL interested farmers — assigned and unassigned both
    all_interested = (
        Team1Assignment.objects
        .filter(interest_status='interested')
        .select_related('farmer', 'assigned_to', 'team2_followup__assigned_to')
    )

    team2_users = User.objects.filter(role='team2')
    team2_summary = []
    for u in team2_users:
        assignments = Team2Assignment.objects.filter(assigned_to=u).select_related('farmer')
        team2_summary.append({
            'user': u,
            'total': assignments.count(),
            'contacted': assignments.filter(followup_status='contacted').count(),
            'pending': assignments.filter(followup_status='pending').count(),
            'confirmed': assignments.filter(final_response='confirmed').count(),
            'assignments': assignments,
        })

    unassigned_count = all_interested.filter(team2_followup__isnull=True).count()
    assigned_count   = all_interested.filter(team2_followup__isnull=False).count()

    return render(request, 'core/assign_team2.html', {
        'all_interested': all_interested,
        'interested_assignments': all_interested.filter(team2_followup__isnull=True),
        'form': form,
        'team2_summary': team2_summary,
        'team2_users': team2_users,
        'total_interested_unassigned': unassigned_count,
        'total_interested_assigned': assigned_count,
    })


@role_required('admin')
def user_create_team2(request):
    """Quick create a Team 2 member from the assign page modal."""
    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        password   = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        phone      = request.POST.get('phone', '').strip()
        if username and password:
            if User.objects.filter(username=username).exists():
                messages.error(request, f"Username '{username}' already exists.")
            else:
                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name, last_name=last_name,
                    role='team2'
                )
                user.phone = phone
                user.save()
                messages.success(request, f"Team 2 member '{first_name} {last_name}' added! Login: {username}")
        else:
            messages.error(request, "Username and password are required.")
    return redirect('assign_team2')


@role_required('admin', 'manager')
def interested_farmers_view(request):
    """Shows all farmers with interest status filter + charts."""
    search = request.GET.get('search', '')
    interest_filter = request.GET.get('interest_filter', '')

    # All team1 assignments (not just interested)
    all_assignments = Team1Assignment.objects.select_related('farmer', 'assigned_to').all()

    if search:
        all_assignments = all_assignments.filter(
            Q(farmer__name__icontains=search) |
            Q(farmer__phone__icontains=search) |
            Q(farmer__village__icontains=search)
        )
    if interest_filter:
        all_assignments = all_assignments.filter(interest_status=interest_filter)

    # Chart data counts
    total_interested = Team1Assignment.objects.filter(interest_status='interested').count()
    not_interested   = Team1Assignment.objects.filter(interest_status='not_interested').count()
    no_response      = Team1Assignment.objects.filter(interest_status='no_response').count()
    not_called       = Team1Assignment.objects.filter(interest_status='not_called').count()

    # Per-agent stats for charts
    team1_users = User.objects.filter(role='team1')
    agent_stats = []
    for u in team1_users:
        total = Team1Assignment.objects.filter(assigned_to=u).count()
        intr  = Team1Assignment.objects.filter(assigned_to=u, interest_status='interested').count()
        rate  = round((intr / total * 100), 1) if total > 0 else 0
        agent_stats.append({
            'name': u.get_full_name() or u.username,
            'total': total,
            'interested': intr,
            'rate': rate,
        })

    return render(request, 'core/interested_farmers.html', {
        'all_assignments': all_assignments,
        'search': search,
        'interest_filter': interest_filter,
        'total_interested': total_interested,
        'not_interested': not_interested,
        'no_response': no_response,
        'not_called': not_called,
        'agent_stats': agent_stats,
    })


@role_required('admin', 'manager')
def confirmed_farmers_view(request):
    """Shows all confirmed/ready-to-work farmers."""
    search = request.GET.get('search', '')
    confirmed_qs = Team2Assignment.objects.select_related(
        'farmer', 'assigned_to', 'team1_assignment__assigned_to'
    ).all()
    if search:
        confirmed_qs = confirmed_qs.filter(
            Q(farmer__name__icontains=search) | Q(farmer__phone__icontains=search)
        )
    total_confirmed = Team2Assignment.objects.count()
    contacted = Team2Assignment.objects.filter(followup_status='contacted').count()
    pending = Team2Assignment.objects.filter(followup_status='pending').count()
    declined = Team2Assignment.objects.filter(final_response='declined').count()
    return render(request, 'core/confirmed_farmers.html', {
        'confirmed_farmers': confirmed_qs,
        'search': search,
        'total_confirmed': total_confirmed,
        'contacted': contacted,
        'pending': pending,
        'declined': declined,
    })


@role_required('admin', 'manager')
def ready_to_work_view(request):
    """Shows only farmers confirmed as Ready to Work (final_response='confirmed')."""
    search = request.GET.get('search', '')
    qs = Team2Assignment.objects.filter(
        final_response='confirmed'
    ).select_related('farmer', 'assigned_to', 'team1_assignment__assigned_to')

    if search:
        qs = qs.filter(
            Q(farmer__name__icontains=search) |
            Q(farmer__phone__icontains=search) |
            Q(farmer__village__icontains=search)
        )

    unique_villages = qs.values('farmer__village').distinct().count()
    unique_crops = qs.values('farmer__crop').distinct().count()
    unique_agents = qs.values('assigned_to').distinct().count()

    return render(request, 'core/ready_to_work.html', {
        'ready_farmers': qs,
        'total': qs.count(),
        'unique_villages': unique_villages,
        'unique_crops': unique_crops,
        'unique_agents': unique_agents,
        'search': search,
    })


@role_required('admin')
def all_farmers_view(request):
    """Admin sees all farmers with current statuses."""
    search = request.GET.get('search', '')
    t1_status = request.GET.get('t1_status', '')
    interest = request.GET.get('interest', '')

    assignments = Team1Assignment.objects.select_related('farmer', 'assigned_to').all()
    if search:
        assignments = assignments.filter(
            Q(farmer__name__icontains=search) | Q(farmer__phone__icontains=search) |
            Q(farmer__village__icontains=search)
        )
    if t1_status:
        assignments = assignments.filter(call_status=t1_status)
    if interest:
        assignments = assignments.filter(interest_status=interest)

    return render(request, 'core/all_farmers.html', {
        'assignments': assignments,
        'search': search,
        't1_status': t1_status,
        'interest': interest,
    })


# ─────────────────────────────────────────────
#  TEAM 1 VIEWS
# ─────────────────────────────────────────────

@role_required('team1')
def team1_dashboard(request):
    """Team 1 member sees their assigned farmers."""
    assignments = (
        Team1Assignment.objects
        .filter(assigned_to=request.user)
        .select_related('farmer')
    )
    search = request.GET.get('search', '')
    filter_val = request.GET.get('filter', '')

    if search:
        assignments = assignments.filter(
            Q(farmer__name__icontains=search) | Q(farmer__phone__icontains=search) |
            Q(farmer__village__icontains=search)
        )
    if filter_val == 'interested':
        assignments = assignments.filter(interest_status='interested')
    elif filter_val == 'pending':
        assignments = assignments.filter(call_status='pending')
    elif filter_val == 'called':
        assignments = assignments.filter(call_status='completed')
    elif filter_val == 'not_interested':
        assignments = assignments.filter(interest_status='not_interested')
    elif filter_val == 'no_response':
        assignments = assignments.filter(interest_status='no_response')

    total = Team1Assignment.objects.filter(assigned_to=request.user).count()
    completed = Team1Assignment.objects.filter(assigned_to=request.user, call_status='completed').count()
    pending = total - completed
    interested = Team1Assignment.objects.filter(assigned_to=request.user, interest_status='interested').count()

    return render(request, 'core/team1_dashboard.html', {
        'assignments': assignments,
        'total': total,
        'completed': completed,
        'pending': pending,
        'interested': interested,
        'search': search,
    })


@role_required('team1')
def team1_quick_update(request, pk):
    """Quick inline interest status update from dropdown."""
    assignment = get_object_or_404(Team1Assignment, pk=pk, assigned_to=request.user)
    if request.method == 'POST':
        interest = request.POST.get('interest_status', '')
        if interest:
            assignment.interest_status = interest
            if interest != 'not_called':
                assignment.call_status = 'completed'
            assignment.save()
    return redirect('team1_dashboard')


@role_required('team1')
def team1_update(request, pk):
    """Team 1 updates call/interest status for a farmer."""
    assignment = get_object_or_404(Team1Assignment, pk=pk, assigned_to=request.user)
    form = Team1UpdateForm(request.POST or None, instance=assignment)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Updated status for {assignment.farmer.name}.")
        return redirect('team1_dashboard')
    return render(request, 'core/team1_update.html', {
        'form': form,
        'assignment': assignment,
        'farmer': assignment.farmer,
    })


# ─────────────────────────────────────────────
#  TEAM 2 VIEWS
# ─────────────────────────────────────────────

@role_required('team2')
def team2_dashboard(request):
    """Team 2 member sees their follow-up farmers."""
    assignments = (
        Team2Assignment.objects
        .filter(assigned_to=request.user)
        .select_related('farmer', 'team1_assignment')
    )
    search = request.GET.get('search', '')
    filter_val = request.GET.get('filter', '')

    if search:
        assignments = assignments.filter(
            Q(farmer__name__icontains=search) | Q(farmer__phone__icontains=search)
        )
    if filter_val == 'contacted':
        assignments = assignments.filter(followup_status='contacted')
    elif filter_val == 'pending':
        assignments = assignments.filter(followup_status='pending')
    elif filter_val == 'confirmed':
        assignments = assignments.filter(final_response='confirmed')
    elif filter_val == 'declined':
        assignments = assignments.filter(final_response='declined')
    elif filter_val == 'callback':
        assignments = assignments.filter(final_response='callback')
    elif filter_val == 'not_reachable':
        assignments = assignments.filter(followup_status='not_reachable')

    total = Team2Assignment.objects.filter(assigned_to=request.user).count()
    contacted = Team2Assignment.objects.filter(assigned_to=request.user, followup_status='contacted').count()
    pending = total - contacted
    confirmed = Team2Assignment.objects.filter(assigned_to=request.user, final_response='confirmed').count()

    return render(request, 'core/team2_dashboard.html', {
        'assignments': assignments,
        'total': total,
        'contacted': contacted,
        'pending': pending,
        'confirmed': confirmed,
        'search': search,
    })


@role_required('team2')
def team2_quick_reached(request, pk):
    """Quick update for Reached / Not Reached."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team2Assignment, pk=pk, assigned_to=request.user)
        status = request.POST.get('followup_status', 'pending')
        remarks = request.POST.get('followup_remarks', '')
        if status in ['pending', 'contacted', 'not_reachable']:
            assignment.followup_status = status
            if remarks:
                assignment.followup_remarks = remarks
            assignment.save(update_fields=['followup_status', 'followup_remarks', 'updated_at'])
    next_page = request.POST.get('next', '')
    if next_page == 'update':
        return redirect('team2_update', pk=pk)
    return redirect('team2_dashboard')


@role_required('team2')
def team2_quick_decision(request, pk):
    """Quick update for Ready to Work / Decline with optional reason."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team2Assignment, pk=pk, assigned_to=request.user)
        response = request.POST.get('final_response', 'awaiting')
        remarks = request.POST.get('followup_remarks', '')
        if response in ['awaiting', 'confirmed', 'declined']:
            assignment.final_response = response
            if remarks:
                assignment.followup_remarks = remarks
            assignment.save(update_fields=['final_response', 'followup_remarks', 'updated_at'])
    next_page = request.POST.get('next', '')
    if next_page == 'update':
        return redirect('team2_update', pk=pk)
    return redirect('team2_dashboard')



@role_required('team2')
def team2_update(request, pk):
    """Team 2 updates follow-up status."""
    assignment = get_object_or_404(Team2Assignment, pk=pk, assigned_to=request.user)
    form = Team2UpdateForm(request.POST or None, instance=assignment)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Follow-up updated for {assignment.farmer.name}.")
        return redirect('team2_dashboard')
    return render(request, 'core/team2_update.html', {
        'form': form,
        'assignment': assignment,
        'farmer': assignment.farmer,
    })


# ─────────────────────────────────────────────
#  MANAGER VIEWS
# ─────────────────────────────────────────────

@role_required('manager', 'admin')
def manager_dashboard(request):
    """Full funnel analytics for managers — mirrors CSV export."""
    total = Farmer.objects.count()
    assigned_t1 = Team1Assignment.objects.count()
    called = Team1Assignment.objects.filter(call_status='completed').count()
    interested = Team1Assignment.objects.filter(interest_status='interested').count()
    not_interested = Team1Assignment.objects.filter(interest_status='not_interested').count()
    no_response_t1 = Team1Assignment.objects.filter(interest_status='no_response').count()
    assigned_t2 = Team2Assignment.objects.count()
    contacted_t2 = Team2Assignment.objects.filter(followup_status='contacted').count()
    not_reached = Team2Assignment.objects.filter(followup_status='not_reachable').count()
    confirmed = Team2Assignment.objects.filter(final_response='confirmed').count()
    declined = Team2Assignment.objects.filter(final_response='declined').count()
    awaiting = Team2Assignment.objects.filter(final_response='awaiting').count()

    team1_leaderboard = (
        Team1Assignment.objects
        .values('assigned_to__first_name', 'assigned_to__last_name', 'assigned_to__username')
        .annotate(
            total=Count('id'),
            called=Count('id', filter=Q(call_status='completed')),
            interested=Count('id', filter=Q(interest_status='interested')),
        ).order_by('-interested')
    )

    team2_leaderboard = (
        Team2Assignment.objects
        .values('assigned_to__first_name', 'assigned_to__last_name', 'assigned_to__username')
        .annotate(
            total=Count('id'),
            contacted=Count('id', filter=Q(followup_status='contacted')),
            confirmed=Count('id', filter=Q(final_response='confirmed')),
        ).order_by('-confirmed')
    )

    # Full data table — same columns as CSV export
    search = request.GET.get('search', '')
    interest_filter = request.GET.get('interest', '')
    response_filter = request.GET.get('response', '')

    all_farmers_qs = Farmer.objects.select_related(
        'team1_assignment__assigned_to',
        'team2_assignment__assigned_to',
    ).all().order_by('name')

    if search:
        all_farmers_qs = all_farmers_qs.filter(
            Q(name__icontains=search) | Q(phone__icontains=search) |
            Q(village__icontains=search) | Q(district__icontains=search)
        )
    if interest_filter:
        all_farmers_qs = all_farmers_qs.filter(
            team1_assignment__interest_status=interest_filter
        )
    if response_filter:
        all_farmers_qs = all_farmers_qs.filter(
            team2_assignment__final_response=response_filter
        )

    return render(request, 'core/manager_dashboard.html', {
        'total': total,
        'assigned_t1': assigned_t1,
        'called': called,
        'interested': interested,
        'not_interested': not_interested,
        'no_response_t1': no_response_t1,
        'assigned_t2': assigned_t2,
        'contacted_t2': contacted_t2,
        'not_reached': not_reached,
        'confirmed': confirmed,
        'declined': declined,
        'awaiting': awaiting,
        'team1_leaderboard': team1_leaderboard,
        'team2_leaderboard': team2_leaderboard,
        'all_farmers': all_farmers_qs,
        'search': search,
        'interest_filter': interest_filter,
        'response_filter': response_filter,
    })



# ─────────────────────────────────────────────
#  EXPORT
# ─────────────────────────────────────────────

@role_required('admin', 'manager')
def export_farmers_csv(request):
    """Export all farmer data as CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="farmers_export.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Farmer ID', 'Name', 'Phone', 'Village', 'District', 'Crop',
        'Team 1 Agent', 'Call Status', 'Interest Status',
        'Team 2 Agent', 'Followup Status', 'Final Response'
    ])
    for farmer in Farmer.objects.select_related(
        'team1_assignment__assigned_to',
        'team2_assignment__assigned_to'
    ).all():
        t1 = getattr(farmer, 'team1_assignment', None)
        t2 = getattr(farmer, 'team2_assignment', None)
        writer.writerow([
            farmer.farmer_id, farmer.name, farmer.phone,
            farmer.village, farmer.district, farmer.crop,
            str(t1.assigned_to) if t1 and t1.assigned_to else '',
            t1.get_call_status_display() if t1 else '',
            t1.get_interest_status_display() if t1 else '',
            str(t2.assigned_to) if t2 and t2.assigned_to else '',
            t2.get_followup_status_display() if t2 else '',
            t2.get_final_response_display() if t2 else '',
        ])
    return response


    team1_users = list(User.objects.filter(role='team1'))
    if not team1_users:
        messages.error(request, "No Team 1 members found. Add members first.")
        return redirect('assign_team1')

    unassigned = list(Farmer.objects.filter(team1_assignment__isnull=True))
    if not unassigned:
        messages.info(request, "All farmers are already assigned.")
        return redirect('assign_team1')

    count = 0
    for i, farmer in enumerate(unassigned):
        agent = team1_users[i % len(team1_users)]
        Team1Assignment.objects.get_or_create(
            farmer=farmer,
            defaults={'assigned_to': agent, 'assigned_by': request.user}
        )
        count += 1

    messages.success(request, f"{count} farmers auto-assigned equally among {len(team1_users)} Team 1 members.")
    return redirect('assign_team1')


@role_required('admin')
def unassign_team1(request, pk):
    """Admin removes a farmer from a Team 1 assignment — farmer goes back to unassigned pool."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team1Assignment, pk=pk)
        farmer_name = assignment.farmer.name
        agent_name = assignment.assigned_to.get_full_name() if assignment.assigned_to else 'agent'
        assignment.delete()
        messages.success(request, f"{farmer_name} has been unassigned from {agent_name} and is back in the unassigned pool.")
    return redirect('assign_team1')


@role_required('admin')
def unassign_team2(request, pk):
    """Admin removes a farmer from a Team 2 assignment — farmer goes back to interested pool."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team2Assignment, pk=pk)
        farmer_name = assignment.farmer.name
        agent_name = assignment.assigned_to.get_full_name() if assignment.assigned_to else 'agent'
        assignment.delete()
        messages.success(request, f"{farmer_name} has been unassigned from {agent_name} and is back in the interested pool.")
    return redirect('assign_team2')


# ─────────────────────────────────────────────
#  UNASSIGN VIEWS
# ─────────────────────────────────────────────

@role_required('admin')
def unassign_team1(request, pk):
    """Admin removes a farmer from Team 1 — farmer goes back to unassigned pool."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team1Assignment, pk=pk)
        farmer_name = assignment.farmer.name
        agent_name = assignment.assigned_to.get_full_name() if assignment.assigned_to else 'agent'
        assignment.delete()
        messages.success(request, f"{farmer_name} has been unassigned from {agent_name} and is back in the unassigned pool.")
    return redirect('assign_team1')


@role_required('admin')
def unassign_team2(request, pk):
    """Admin removes a farmer from Team 2 — farmer goes back to interested pool."""
    if request.method == 'POST':
        assignment = get_object_or_404(Team2Assignment, pk=pk)
        farmer_name = assignment.farmer.name
        agent_name = assignment.assigned_to.get_full_name() if assignment.assigned_to else 'agent'
        assignment.delete()
        messages.success(request, f"{farmer_name} has been unassigned from {agent_name} and is back in the interested pool.")
    return redirect('assign_team2')
